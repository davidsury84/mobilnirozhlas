import sys,json
from openpyxl import load_workbook
from collections import defaultdict
import datetime

SKNAME={'324':'Plošiny a manipulační technika','255':'Sběrače dešťové vody','147':'Nádoby na zimní posyp',
'141':'Venkovní odpadkové koše','257':'Podzemní nádrže','144':'Plastové kontejnery 1100 l',
'317':'Interiérové koše na tříděný odpad','353':'Lavičky a mobiliář','320':'Kontejnery na textil',
'265':'Plastové palety','326':'Zahrada a domácnost','369':'Přístřešky na kontejnery',
'366':'Záchytné vany pod IBC','371':'Skladovací kontejnery FCM','391':'Rohože pro zpevnění štěrku',
'100':'Sloupky/kotvení','295':'Kontejnery CLE-CO','278':'Kompostéry','152':'Sklolaminátové kontejnery',
'313':'Nádoby na odpad','297':'Kontejnery CLE-PG','282':'Vodoměrné šachty','402':'Koše na tříděný odpad'}

def load(fn):
    wb=load_workbook(fn,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); H=[str(c).strip() if c is not None else '' for c in next(it)]
    ix={k:H.index(k) for k in H if k}
    out=[]
    for r in it:
        try: cc=float(r[ix['CC bez daní']] or 0)
        except: cc=0
        try: y=int(r[ix['Datum případu (R)']]); m=int(r[ix['Datum případu (M)']])
        except: continue
        d=r[ix['Datum případu']]
        wd=None
        if isinstance(d,datetime.datetime) or isinstance(d,datetime.date): wd=d.weekday()
        cust=(str(r[ix['Název']] or '').strip() or str(r[ix['Příjmení']] or '').strip() or str(r[ix['Č. org.']] or '').strip())
        cname=(str(r[ix['Název']] or '').strip() or str(r[ix['Příjmení']] or '').strip())
        out.append({'sk':str(r[ix['SK']]).strip() if r[ix['SK']] is not None else '','reg':str(r[ix['Reg. č.']] or '').strip(),
            'nazev':str(r[ix['Název 1']] or '').strip(),'cc':cc,'y':y,'m':m,'wd':wd,
            'ord':str(r[ix['Číslo zakázky']] or ''),'cust':cust,'cname':cname})
    return out

def catname(sk,names):
    if sk in SKNAME: return SKNAME[sk]
    top=sorted(names[sk].items(),key=lambda x:-x[1])[0][0] if names.get(sk) else ('SK '+sk)
    return top[:34]

def build(rows):
    ym=defaultdict(float); ymo=defaultdict(set)
    yr=defaultdict(float); yro=defaultdict(set); yrc=defaultdict(set)
    h1=defaultdict(float); skrev=defaultdict(float); prodrev=defaultdict(float)
    names=defaultdict(lambda: defaultdict(float))
    skm=defaultdict(lambda: defaultdict(float))     # sk → měsíc → rev (sezónnost kategorie)
    h1sk=defaultdict(lambda: defaultdict(float))
    wd=[0]*7; wdseen=set()
    custrev=defaultdict(float); custord=defaultdict(set); custname={}
    custyears=defaultdict(set); yrev_by_cust=defaultdict(lambda: defaultdict(float))
    allord=set(); skus=set(); ordrev=defaultdict(float)
    for r in rows:
        k=f"{r['y']}-{r['m']:02d}"
        ym[k]+=r['cc']; ymo[k].add(r['ord'])
        yr[r['y']]+=r['cc']; yro[r['y']].add(r['ord']); yrc[r['y']].add(r['cust'])
        if r['m']<=6: h1[r['y']]+=r['cc']; h1sk[r['sk']][r['y']]+=r['cc']
        skrev[r['sk']]+=r['cc']; prodrev[r['nazev']]+=r['cc']; names[r['sk']][r['nazev']]+=r['cc']
        if 2021<=r['y']<=2025: skm[r['sk']][r['m']]+=r['cc']
        if r['wd'] is not None and r['ord'] not in wdseen: wdseen.add(r['ord']); wd[r['wd']]+=1
        custrev[r['cust']]+=r['cc']; custord[r['cust']].add(r['ord'])
        if r['cname']: custname[r['cust']]=r['cname']
        custyears[r['cust']].add(r['y']); yrev_by_cust[r['y']][r['cust']]+=r['cc']
        allord.add(r['ord']); skus.add(r['sk']+'-'+r['reg']); ordrev[r['ord']]+=r['cc']
    years=sorted(yr)
    mlabels=sorted(ym)
    monthly={'labels':mlabels,'rev':[round(ym[k],1) for k in mlabels],'orders':[len(ymo[k]) for k in mlabels],
             'aov':[round(ym[k]/len(ymo[k]),1) if ymo[k] else 0 for k in mlabels]}
    yearly={'labels':years,'rev':[round(yr[y],1) for y in years],'orders':[len(yro[y]) for y in years],
            'cust':[len(yrc[y]) for y in years],'aov':[round(yr[y]/len(yro[y]),1) if yro[y] else 0 for y in years]}
    H1={'labels':years,'rev':[round(h1[y],1) for y in years]}
    # sezónní index 2021-2025
    base=[y for y in years if 2021<=y<=2025]
    mtot=[0.0]*12
    for r in rows:
        if r['y'] in base: mtot[r['m']-1]+=r['cc']
    avg=sum(mtot)/12 or 1
    season=[round(v/avg*100,1) for v in mtot]
    def catseason(sk):
        tot=sum(skm[sk].values()) or 1
        return [round(skm[sk].get(m+1,0)/tot*100,1) for m in range(12)]
    topsk_list=sorted(skrev.items(),key=lambda x:-x[1])[:10]
    topsk={'labels':[catname(sk,names) for sk,_ in topsk_list],'rev':[round(v,1) for _,v in topsk_list]}
    tp=sorted(prodrev.items(),key=lambda x:-x[1])[:12]
    topprod={'labels':[n[:46] for n,_ in tp],'rev':[round(v,1) for _,v in tp]}
    diff={sk:(h1sk[sk].get(max(years),0)-h1sk[sk].get(max(years)-1,0)) for sk in h1sk}
    g=sorted(diff.items(),key=lambda x:-x[1])[:6]; l=sorted(diff.items(),key=lambda x:x[1])[:6]
    gain={'labels':[catname(sk,names) for sk,_ in g],'d':[round(v,1) for _,v in g]}
    lose={'labels':[catname(sk,names) for sk,_ in l],'d':[round(v,1) for _,v in l]}
    # retence bez 2 velkoodběratelů
    big=set([c for c,_ in sorted(custrev.items(),key=lambda x:-x[1])[:2]])
    firstyear={}
    for c in custyears: firstyear[c]=min(custyears[c])
    rs=[]
    for y in years:
        tot=sum(v for c,v in yrev_by_cust[y].items() if c not in big) or 1
        yo=defaultdict(set)
        for rr in rows:
            if rr['y']==y: yo[rr['cust']].add(rr['ord'])
        ret=sum(v for c,v in yrev_by_cust[y].items() if c not in big and (firstyear.get(c,y)<y or len(yo[c])>1))
        rs.append(round(ret/tot*100,1))
    retshare={'labels':years,'share':rs}
    tc=sorted(custrev.items(),key=lambda x:-x[1])[:10]
    topcust={'labels':[custname.get(c,c)[:34] for c,_ in tc],'rev':[round(v,1) for _,v in tc],'orders':[len(custord[c]) for c,_ in tc]}
    aovs=sorted(ordrev.values()); med=aovs[len(aovs)//2] if aovs else 0
    rep=sum(1 for c in custord if len(custord[c])>1)
    reprev=sum(custrev[c] for c in custord if len(custord[c])>1)
    totrev=sum(yr.values())
    b2b=sum(v for c,v in custrev.items() if custname.get(c,'') and any(t in custname.get(c,'') for t in ['s.r.o','a.s','spol','sp. z','GmbH','s. r. o','o.p.s','z.s','obec','Obec','Město','město']))
    stats={'total_rev':round(totrev,1),'orders':len(allord),'customers':len(custrev),'skus':len(skus),
           'aov_median':round(med,1),'repeat_pct':round(rep/len(custrev)*100,1) if custrev else 0,
           'repeat_rev_pct':round(reprev/totrev*100,1) if totrev else 0,'b2b_rev_pct':round(b2b/totrev*100,1) if totrev else 0}
    return {'monthly':monthly,'yearly':yearly,'h1':H1,'season':season,'season_posyp':catseason('147'),
            'season_voda':catseason('255'),'topsk':topsk,'topprod':topprod,'gain':gain,'lose':lose,
            'retshare':retshare,'weekday':wd,'topcust':topcust,'stats':stats}

if __name__=='__main__':
    rows=load(sys.argv[1])
    d=build(rows)
    print(json.dumps(d,ensure_ascii=False))
